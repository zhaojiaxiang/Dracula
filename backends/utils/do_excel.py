# from openpyxl.workbook import Workbook
# from openpyxl.styles import Font, Fill
# wb = Workbook()
# ws = wb.active
#
#
# col = ws.column_dimensions['A']
# col.font = Font(bold=True)
# row = ws.row_dimensions[1]
# row.font = Font(underline="single")
# ws.append([1,2,3,4,5])
#
# wb.save("test.xlsx")
# # wb.save("test.xlsx")

from openpyxl import Workbook, load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.table import Table, TableStyleInfo

# wb = Workbook()
# ws = wb.active

# data = [
#     ['Apples', 10000, 5000, 8000, 6000],
#     ['Pears',   2000, 3000, 4000, 5000],
#     ['Bananas', 6000, 6000, 6500, 6000],
#     ['Oranges',  500,  300,  200,  700],
# ]
# #
# # add column headings. NB. these must be strings
# ws.append(["Fruit", "2011", "2012", "2013", "2014"])
# for row in data:
#     ws.append(row)
#
# tab = Table(displayName="Table1", ref="A1:E5")
#
# # Add a default style with striped rows and banded columns
# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
# tab.tableStyleInfo = style
# ws.add_table(tab)
# wb.save("table.xlsx")



# style = Font(bold=True)
# for i in range(1, 100000):
#     for x in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
#         ws[x+str(i)] = x+str(i)
#         ws[x + str(i)].font = style

# from openpyxl.styles import numbers
#
# # use pre-defined values
# ws.cell['A1'].number_format = numbers.FORMAT_GENERAL
# ws.cell(row=2, column=4).number_format = numbers.FORMAT_DATE_XLSX15
#
# # use strings
# ws.cell['A2'].number_format = 'General'
# ws.cell(row=3, column=5).number_format = 'd-mmm-yy'
# ws.cell['A3'].number_format = '0.00'
# ws.cell['A4'].number_format = '0.00%'
# ws.cell['A5'].number_format = '_ * #,##0_ ;_ * -#,##0_ ;_ * "-"??_ ;_ @_ '

# 0.00_ ;[Red]\-0.00\
#
# wb = load_workbook('styled.xlsx')
# sheet = wb["Sheet"]
# A1 = sheet.cell(1, 1)
# print(A1.number_format)


# red_text = Font(color="FF0000")
# dxf = DifferentialStyle(font=red_text)
# rule = Rule(type="containsText", operator="containsText", text="-", dxf=dxf)
# rule.formula = ['NOT(ISERROR(SEARCH("-",A1)))']
# ws.conditional_formatting.add('A1:F40', rule)
# wb.save("test.xlsx")


# ws["A1:A2"].value = 1
# wb.save("test.xlsx")

# ws.column_dimensions['A:Z'].number_format='#0.00_ ;[Red]-0.00'
# wb.save("test.xlsx")


# ws['A1'] = -1234
# ws['A1'].number_format = '#0.00_ ;[Red]-0.00'
# wb.save("test.xlsx")

# wb = load_workbook("test1.xlsx")
# ws = wb.active
# for a in ws.conditional_formatting:
#     print(a)

# wb = Workbook()
# ws = wb.active
#
# ws.append([1, 2])

# wb1 = Workbook()
# ws1 = wb1.create_sheet()
# ws1.append([22,23,244422,])
# wb1.save('test2.xlsx')
#
# wb2 = Workbook()
# ws2 = wb1.create_sheet()
# ws2.append([22,23,2555522,])
# wb2.save('test2.xlsx')
#
# merge_Excel('test2.xlsx','test2.xlsx')

wb1 = Workbook()
ws1 = wb1.active

for i in range(1000001):
    ws1.append([1111111, 2111111, 3111111, 4111111, 5111111, 6111111, 7111111])

wb1.save('t.xlsx')



